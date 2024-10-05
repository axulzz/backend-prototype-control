import openpyxl
from io import BytesIO

from rest_framework.viewsets import ReadOnlyModelViewSet
from drf_excel.renderers import XLSXRenderer
from drf_excel.mixins import XLSXFileMixin

from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework.renderers import BaseRenderer
from openpyxl.utils import get_column_letter
from apps.cores.models import AcademicGroup, TypeInvestigation
from apps.school.models import Teacher


class CustomXLSXRenderer(XLSXRenderer):
    def render(self, data, media_type=None, renderer_context=None):
        response = renderer_context["response"]
        view = renderer_context["view"]

        wb = openpyxl.Workbook()
        ws = wb.active

        # Set the filename from the view
        filename = getattr(view, "filename", "export.xlsx")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Get column headers from the view
        column_header = getattr(view, "column_header", {})
        column_titles = column_header.get("titles", [])

        # Write column headers to the first row
        for col_num, title in enumerate(column_titles, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = Font(bold=True)

        # Save the workbook to the response
        output = BytesIO()
        wb.save(output)

        return output.getvalue()

    def _add_data_validation(self, ws, titles, header_name, choices):
        # Find the column index for the specified header
        column_index = next(
            (i + 1 for i, title in enumerate(titles) if title == header_name), None
        )
        if column_index is not None:
            column_letter = get_column_letter(column_index)
            # Prepare the data validation with the list of choices
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(choices)}"',
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
            )
            # Apply validation to all rows in that column
            validation.add(f"{column_letter}2:{column_letter}10000")
            ws.add_data_validation(validation)


class CustomStudentXLSRender(CustomXLSXRenderer):
    TURN = ["MATUTINO", "VESPERTINO"]
    SPECIALITY = [
        "TECNICO EN PROGRAMACION",
        "TECNICO EN CONTABILIDAD",
        "TECNICO EN SECRETARIADO EJECUTIVO BILINGUE",
        "TECNICO EN CIENCIA DE DATOS E INFORMACION",
    ]

    @classmethod
    def get_groups(cls):
        return [group.text for group in AcademicGroup.objects.all()]

    def render(self, data, media_type=None, renderer_context=None):
        response = renderer_context["response"]
        view = renderer_context["view"]

        wb = openpyxl.Workbook()
        ws = wb.active

        # Set the filename from the view
        filename = getattr(view, "filename", "export.xlsx")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Get column headers from the view
        column_header = getattr(view, "column_header", {})
        column_titles = column_header.get("titles", [])

        # Write column headers to the first row
        for col_num, title in enumerate(column_titles, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = Font(bold=True)

        self._add_data_validation(ws, column_titles, "Especialidad", self.SPECIALITY)
        self._add_data_validation(ws, column_titles, "Turno", self.TURN)
        self._add_data_validation(ws, column_titles, "Grupo", self.get_groups())

        # Save the workbook to the response
        output = BytesIO()
        wb.save(output)

        return output.getvalue()


class CustomPrototypeXLSXRenderer(CustomXLSXRenderer):
    @classmethod
    def get_type_investigation(cls):
        return [
            type_investigation.text
            for type_investigation in TypeInvestigation.objects.all()
        ]

    MODALITY = [
        "TECNOLOGICO",
        "SOFTWARE",
        "DIDACTICO",
        "EMPRENDEDOR_VERDE",
        "EMPRENDEDOR_SOCIAL",
        "EMPRENDEDOR_TECNOLOGICO",
    ]

    def render(self, data, media_type=None, renderer_context=None):
        response = renderer_context["response"]
        view = renderer_context["view"]

        wb = openpyxl.Workbook()
        ws = wb.active

        # Set the filename from the view
        filename = getattr(view, "filename", "export.xlsx")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Get column headers from the view
        column_header = getattr(view, "column_header", {})
        column_titles = column_header.get("titles", [])

        # Write column headers to the first row
        for col_num, title in enumerate(column_titles, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = Font(bold=True)

        self._add_data_validation(
            ws, column_titles, "Línea de investigación", self.get_type_investigation()
        )
        self._add_data_validation(ws, column_titles, "Modalidad", self.MODALITY)

        self._add_data_validation(ws, column_titles, "Autor 1", ["1", " "])
        self._add_data_validation(ws, column_titles, "Autor 2", ["2", " "])
        self._add_data_validation(ws, column_titles, "Autor 3", ["3", " "])
        self._add_data_validation(ws, column_titles, "Autor 4", ["4", " "])

        # Save the workbook to the response
        output = BytesIO()
        wb.save(output)

        return output.getvalue()


class TemplateBase(XLSXFileMixin, ReadOnlyModelViewSet):
    renderer_classes = (CustomXLSXRenderer,)


class TeacherTemplate(TemplateBase):
    filename = "PlantillaMaestros.xlsx"
    column_header = {"titles": []}


class StudentTamplete(TemplateBase):
    filename = "PlantillaAlumnos.xlsx"
    renderer_classes = (CustomStudentXLSRender,)

    column_header = {
        "titles": [
            "Nombres",
            "Apellidos",
            "CURP",
            "Dirección",
            "Correo institucional",
            "Numero celular",
            "Turno",
            "Numero de control escolar",
            "Grupo",
            "Especialidad",
        ]
    }


class PrototypeTemplate(TemplateBase):
    filename = "PlatillasPrototipos.xlsx"
    renderer_classes = (CustomPrototypeXLSXRenderer,)

    column_header = {
        "titles": [
            "Nombre del prototipo",
            "No.registro",
            "Modalidad",
            "Línea de investigación",
            "Maestro(a) de métodos(correo)",
            "Asesor(a) metodologico(correo)",
            "Asesor(a) técnico(correo)",
            "Autor 1",
            "Correo institucional de alumno(a)",
            "Autor 2",
            "Correo institucional de alumno(a)",
            "Autor 3",
            "Correo institucional de alumno(a)",
            "Autor 4",
            "Correo institucional de alumno(a)",
        ],
    }


class PrototypeDonwload(TemplateBase):
    filename = "ReportePrototipos.xlsx"
    column_header = {
        "titles": [
            "NOMBRE PROTOTIPO",
            "NUMERO REGISTRO",
            "Modalidad ",
            "Línea de Investigación",
            "Cal. Preeselección",
            "Maestro de Métodos",
            "ASESOR METODOLOGICO",
            "CLAVE PRESUPUESTAL",
            "NIVEL ACADEMICO ",
            "EMAIL ASESOR",
            "DIRECCION",
            "TELÉFONO",
            "ASESOR TECNICO",
            "CLAVE PRESUPUESTA",
            "NIVEL ACADEMICO",
            "EMAIL ASESOR",
            "DIRECCIÓN",
            "TELÉFONO",
            "Autor",
            "NOMBRE COMPLETO",
            "GRUPO",
            "ESPECIALIDAD",
            "TURNO",
            "No. DE CONTROL ESCOLAR",
            "CORREO INSTITUCIONAL",
            "DIRECCION",
            "TELEFONO",
            "Autor",
            "NOMBRE COMPLETO",
            "GRUPO",
            "ESPECIALIDAD",
            "TURNO",
            "No. DE CONTROL ESCOLAR",
            "CORREO INSTITUCIONAL",
            "DIRECCION",
            "TELEFONO",
            "Autor",
            "NOMBRE COMPLETO",
            "GRUPO",
            "ESPECIALIDAD",
            "TURNO",
            "No. DE CONTROL ESCOLAR",
            "CORREO INSTITUCIONAL",
            "DIRECCION",
            "TELEFONO",
            "Autor",
            "NOMBRE COMPLETO",
            "GRUPO",
            "ESPECIALIDAD",
            "TURNO",
            "No. DE CONTROL ESCOLAR",
            "CORREO INSTITUCIONAL",
            "DIRECCION",
            "TELEFONO",
        ],
    }
