from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
import openpyxl
from faker import Faker
from openpyxl import Workbook

from apps.cores.models import AcademicGroup
from apps.school.models import Student


User = get_user_model()

class Command(BaseCommand):
    help = "Este comando crea maestros"

    def handle(self, *args, **options):
        excel_file_path = "4SEM24.xlsx"
        max_rows = 615
        output_excel_file_path = "generated_passwords.xlsx"

        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = "Passwords"
        output_sheet.append(["Correo", "Contraseña"])
        fake = Faker()
        
        try:
            # Open the Excel file
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
            self.stdout.write(self.style.SUCCESS(f"Archivo {excel_file_path} cargado exitosamente."))

            # Get the active sheet
            sheet = workbook.active

            group = Group.objects.get(name="Alumno").pk

            # Iterate over the rows of the sheet
            for row_number, row in enumerate(sheet.iter_rows(values_only=True, max_row=max_rows)):
                if row_number < 2:  # Skip the header rows
                    continue

                try:
                    if len(row) < 17:
                        self.stdout.write(
                            self.style.ERROR(
                                f"Error en la fila {row_number}: Número incorrecto de valores. Verifica la estructura del archivo Excel."
                            )
                        )
                        continue

                    number_phone = int(row[16]) if row[16] is not None else None

                    turn = row[1]
                    if turn == "vespertino":
                        turn = "T/V"
                    elif turn == "matutino":
                        turn = "T/M"

                    password = fake.password(length=12)
                    user = User.objects.create_user(
                        first_name=row[8],
                        last_name=f"{row[9]} {row[10]}",
                        email=row[5],
                        curp=row[11],
                        turn=turn,
                        number_phone=number_phone,
                        address=f"{row[15]} {row[14]} {row[12]} {row[13]}",
                        password=password,
                    )
                    output_sheet.append([str(row[5]).strip(), password])

                    user.groups.add(group)
                    user.save()

                    modality = row[0]
                    if modality == "PROGRAMACIÓN":
                        modality = "TEP"
                    elif modality == "CONTABILIDAD":
                        modality = "TEC"
                    elif modality == "SECRETARIADO EJECUTIVO BILINGÜE":
                        modality = "TESEB"


                    Student.objects.create(
                        user=user,
                        school_control_number=row[4],
                        specialty=modality,
                        # turn=turn,
                        group=AcademicGroup.objects.get(text=f"{row[3][:-1]}-{row[3][1]}"),
                    )
                    self.stdout.write(self.style.SUCCESS(f"Creado Alumno: {user.first_name}"))

                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f"Error al procesar la fila {row_number}: {str(e)}."
                        )
                    )

            # Save the output workbook after processing all rows
            try:
                output_workbook.save(output_excel_file_path)
                self.stdout.write(self.style.SUCCESS(f"Archivo {output_excel_file_path} guardado exitosamente."))
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f"Error al guardar el archivo {output_excel_file_path}: {str(e)}")
                )

            self.stdout.write(self.style.SUCCESS("Proceso completado exitosamente."))

        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f"Error: No se encontró el archivo {excel_file_path}.")
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f"Error al procesar el archivo Excel: {str(e)}")
            )
