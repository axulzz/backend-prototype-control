from datetime import datetime
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.db import IntegrityError

import openpyxl
from apps.cores.models import AcademicGroup, TypeInvestigation
from apps.prototypes.models import Member, Prototype, TeacherRoles
from apps.school.models import Student, Teacher
from faker import Faker


User = get_user_model()


class Command(BaseCommand):
    count = 1
    max_rows = 354
    help = "Este comando crea maestros"

    def get_name(self, name: str):
        first_name = ""
        last_name = ""

        try:
            name_split = name.strip().split(" ")  # type: ignore
            for index, name_field in enumerate(name_split):
                if index < len(name_split) - 2:
                    first_name = f"{first_name} {name_field}"

                else:
                    last_name = f"{last_name} {name_field}"
            return first_name, last_name
        except AttributeError:
            pass

    def get_turn(self, turn: str):
        if turn in ["Vespertino", "vespertino", "Vespertino"]:
            return "T/V"

        if turn in ["Matutino", "MATUTINO"]:
            return "T/M"

    def get_speciality(self, speciality: str):
        if speciality in ["Programacion", "Programación", "Prpgramación"]:
            return "TEP"

        if speciality == "Contabilidad":
            return "TEC"

        if speciality in [
            "Secretariado Ejecutivo Bilingüe",
            "Secretariado",
            "Secretariado Ejecutivo Bilingüe",
        ]:
            return "TESEB"

    def get_group(self, group: str):
        group_base = group.strip()
        group_split = group_base.split("°")

        if len(group_split) >= 2:
            return AcademicGroup.objects.get(
                text=f"{group_split[0].strip()}-{group_split[1].strip()}"
            )

        return AcademicGroup.objects.get(text=f"{group_base[:-1]}-{group_base[1]}")

    def get_modality(self, modality: str):
        if modality == "Tecnologico":
            return "TEC"

        if modality in ["Software", "Software"]:
            return "SW"

        if modality in ["Didactico", "Didáctico"]:
            return "DC"

        if modality in ["EMPRENDEDOR SOCIAL", "Emprendedor Social"]:
            return "ES"

        if modality in ["EMPRENDEDOR VERDE", "Emprendedor Verde", "Emprendedor VERDE"]:
            return "EV"

        if modality in [
            "EMPRENDEDOR EN TECNOLOGÍAS",
            "Emprendedor Tecnológico",
            "Emprendedor Tecnologico",
        ]:
            return "ET"

    def get_students(self):
        excel_file_path = "prototipos2023.xlsx"

        student_anterior = 0
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        sheet = workbook.active
        members = []
        for index, row_prototype in enumerate(sheet.iter_rows(values_only=True, max_row=self.max_rows)):  # type: ignore #

            if index >= self.count:
                author = int(row_prototype[9])  # type: ignore

                if not author >= student_anterior:
                    student_anterior = 0
                    return members

                if row_prototype[10] is not None:
                    first_name, last_name = self.get_name(row_prototype[10])  # type: ignore

                    try:
                        user, create  = User.objects.get_or_create(
                        first_name=first_name.upper(),
                        last_name=last_name.upper(),
                        number_phone=row_prototype[16],  # type: ignore
                        address=f"{row_prototype[17]}",
                        email=row_prototype[15],
                        password=Faker().password(length=12),
                    )
                    
                        if create:
                            student = Student.objects.create(
                                user=user,
                                group=self.get_group(row_prototype[11]),  # type: ignore
                                specialty=self.get_speciality(row_prototype[12].strip()),  # type: ignore
                                turn=self.get_turn(row_prototype[13]),  # type: ignore
                                school_control_number=row_prototype[14],  # type: ignore
                            )

                            member = Member.objects.create(student=student, author=author)
                            members.append(member.id)
                    except:
                        pass

                    student_anterior = author
                    self.count += 1

    def handle(self, *args, **options):
        excel_file_path = "prototipos2023.xlsx"
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        sheet = workbook.active

        for index, row_prototype in enumerate(sheet.iter_rows(values_only=True, max_row=self.max_rows)):  # type: ignore #
            advisor_tecnical = None
            advisor_method = None

            if index >= 1 and row_prototype[5] is not None:
                students = self.get_students()

                try:
                    first_name_t_method, last_name_t_method = self.get_name(row_prototype[18])  # type: ignore
                except TypeError:
                    first_name_t_method = ""
                    last_name_t_method = ""

                teacher_method = Teacher.objects.filter(
                    user__first_name__icontains=first_name_t_method,
                    user__last_name__icontains=last_name_t_method,
                ).first()

                try:
                    first_name_t_advisor_method, last_name_t_advisor_method = self.get_name(row_prototype[19])  # type: ignore
                except TypeError:
                    first_name_t_advisor_method = ""
                    last_name_t_advisor_method = ""

                teacher_advisor_method = Teacher.objects.filter(
                    user__first_name__icontains=first_name_t_advisor_method,
                    user__last_name__icontains=last_name_t_advisor_method,
                ).first()

                try:
                    first_name_t_advisor_tecnical, last_name_t_advisor_tecnical = self.get_name(row_prototype[25])  # type: ignore
                except TypeError:
                    first_name_t_advisor_tecnical = ""
                    last_name_t_advisor_tecnical = ""

                teacher_advisor_tecnical = Teacher.objects.filter(
                    user__first_name__icontains=first_name_t_advisor_tecnical,
                    user__last_name__icontains=last_name_t_advisor_tecnical,
                ).first()

                if teacher_advisor_method is not None:
                    advisor_method = TeacherRoles.objects.create(
                        teacher_data=teacher_advisor_method, roles="AM"
                    )

                if teacher_advisor_tecnical is not None:
                    advisor_tecnical = TeacherRoles.objects.create(
                        teacher_data=teacher_advisor_tecnical, roles="AT"
                    )

                prototype = Prototype.objects.create(
                    name=row_prototype[5],
                    qualification=row_prototype[8],
                    registry_number=f"{self.get_modality(row_prototype[6])}-{get_index()}",  # type: ignore
                    modality=self.get_modality(row_prototype[6]),  # type: ignore
                    type_investigation=TypeInvestigation.objects.get(
                        text=row_prototype[7].strip(), # type: ignore
                    ),
                    teacher_methods=teacher_method,
                )

                if students:
                    prototype.members.add(*students)
                    prototype.save()

                if advisor_method or advisor_tecnical:
                    prototype.teacher_advisors.add(
                        *(filter(None, [advisor_method, advisor_tecnical]))
                    )
                    prototype.save()

                self.stdout.write(self.style.SUCCESS(f"Creado maestro: {prototype}"))