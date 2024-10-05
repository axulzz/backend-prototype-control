import openpyxl

from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.db import IntegrityError

from apps.cores.models import AcademicGroup
from apps.prototypes.models import Member, Prototype, TeacherRoles
from apps.school.models import Student, Teacher


User = get_user_model()


class Command(BaseCommand):
    count = 1
    max_rows = 354
    help = "Este comando crea maestros"

    def get_name(self, name: str):
        first_name = ""
        last_name = ""

        try:
            name_split = name.strip().split(" ")  # type: ignore
            for index, name_field in enumerate(name_split):
                if index < len(name_split) - 2:
                    first_name = f"{first_name} {name_field}"

                else:
                    last_name = f"{last_name} {name_field}"
            return first_name.upper(), last_name.upper()
        except AttributeError:
            pass

    def get_turn(self, turn: str):
        if turn in ["Vespertino", "vespertino", "Vespertino"]:
            return "T/V"

        if turn in ["Matutino", "MATUTINO"]:
            return "T/M"

    def get_speciality(self, speciality: str):
        if speciality in ["Programacion", "Programación", "Prpgramación"]:
            return "TEP"

        if speciality == "Contabilidad":
            return "TEC"

        if speciality in [
            "Secretariado Ejecutivo Bilingüe",
            "Secretariado",
            "Secretariado Ejecutivo Bilingüe",
        ]:
            return "TESEB"

    def get_group(self, group: str):
        group_base = group.strip()
        group_split = group_base.split("°")

        if len(group_split) >= 2:
            return AcademicGroup.objects.get(
                text=f"{group_split[0].strip()}-{group_split[1].strip()}"
            )

        return AcademicGroup.objects.get(text=f"{group_base[:-1]}-{group_base[1]}")

    def get_modality(self, modality: str):
        if modality in ["Tecnologico", "TECNOLÓGICO"]:
            return "TEC"

        if modality in ["Software", "Software", "SOFTWARE"]:
            return "SW"

        if modality in ["Didactico", "Didáctico", "DIDÁCTICO"]:
            return "DC"

        if modality in ["EMPRENDEDOR SOCIAL", "Emprendedor Social"]:
            return "ES"

        if modality in ["EMPRENDEDOR VERDE", "Emprendedor Verde", "Emprendedor VERDE"]:
            return "EV"

        if modality in [
            "EMPRENDEDOR EN TECNOLOGÍAS",
            "Emprendedor Tecnológico",
            "Emprendedor Tecnologico",
        ]:
            return "ET"

    def get_students(self):
        excel_file_path = "REGISTRO EQUIPOS UNICO.xlsx"
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        sheet = workbook.active
        student_anterior = 0
        members = []

        for index, row_prototype in enumerate(
            sheet.iter_rows(values_only=True, max_row=self.max_rows)
        ):
            if index >= self.count and row_prototype[3] is not None:
                author = int(row_prototype[3])

                if author < student_anterior:
                    student_anterior = 0
                    return members  # Returning members if condition met

                if row_prototype[4] is not None:
                    first_name, last_name = self.get_name(row_prototype[4])

                    try:
                        student = Student.objects.get(
                            user__first_name__icontains=first_name,
                            user__last_name__icontains=last_name,
                        )
                        member = Member.objects.create(
                            student=student, author=row_prototype[3]
                        )
                        members.append(member.id)
                    except Student.DoesNotExist:
                        print(f"Student not found for {first_name} {last_name}")
                    except IntegrityError as e:
                        print(f"Integrity error occurred: {e}")
                    except Exception as e:
                        print(f"An unexpected error occurred: {e}")

                    student_anterior = author
                    self.count += 1

        return members  # Return the final list of member IDs

    def handle(self, *args, **options):
        excel_file_path = "REGISTRO EQUIPOS UNICO.xlsx"
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        sheet = workbook.active

        for index, row_prototype in enumerate(sheet.iter_rows(values_only=True, max_row=self.max_rows)):  # type: ignore #
            advisor_tecnical = None
            advisor_method = None

            if index >= 1 and row_prototype[1] is not None:
                students = self.get_students()

                try:
                    first_name_t_method, last_name_t_method = self.get_name(row_prototype[18])  # type: ignore
                except TypeError:
                    first_name_t_method = ""
                    last_name_t_method = ""

                teacher_method = Teacher.objects.filter(
                    user__first_name__icontains=first_name_t_method,
                    user__last_name__icontains=last_name_t_method,
                ).first()

                if row_prototype[2] is not None:

                    teacher_advisor_method, teacher_advisor_tecnical = get_teachers_by_modality(self.get_modality(str(row_prototype[2])))  # type: ignore

                    if teacher_advisor_method is not None:
                        advisor_method = TeacherRoles.objects.create(
                            teacher_data=teacher_advisor_method, roles="AM"
                        )

                    if teacher_advisor_tecnical is not None:
                        advisor_tecnical = TeacherRoles.objects.create(
                            teacher_data=teacher_advisor_tecnical, roles="AT"
                        )

                    prototype = Prototype.objects.create(
                        name=row_prototype[1],
                        registry_number=f"{self.get_modality(row_prototype[2])}-{get_index()}",  # type: ignore
                        modality=self.get_modality(row_prototype[2]),  # type: ignore
                        teacher_methods=teacher_method,
                    )

                    if students:
                        prototype.members.add(*students)
                        prototype.save()

                    if advisor_method or advisor_tecnical:
                        prototype.teacher_advisors.add(
                            *(filter(None, [advisor_method, advisor_tecnical]))
                        )
                        prototype.save()

                    self.stdout.write(
                        self.style.SUCCESS(f"prototipo: {prototype.name}")
                    )
