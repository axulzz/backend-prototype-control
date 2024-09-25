from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth import get_user_model

import openpyxl
from openpyxl import Workbook
from apps.cores.models import AcademicLevel
from apps.school.models import Teacher
from faker import Faker



User = get_user_model()


class Command(BaseCommand):
    help = "Este comando crea maestros y genera un archivo Excel con las contraseñas y correos"

    def handle(self, *args, **options):
        excel_file_path = "teachers_import.xlsx"
        output_excel_file_path = "generated_passwords_teachers.xlsx"
        max_rows = 72
        error_count = 0

        # Crear un nuevo libro de Excel para almacenar correos y contraseñas
        output_workbook = Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = output_excel_file_path
        output_sheet.append(["Correo", "Contraseña"])

        try:
            # Abre el archivo Excel
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
            group = Group.objects.get(name="Docente").pk

            # Obtiene la hoja activa
            sheet = workbook.active

            # Itera sobre las filas de la hoja
            for row_number, row in enumerate(sheet.iter_rows(values_only=True, max_row=max_rows)):  # type: ignore #
                try:
                    if len(row) >= max_rows:
                        self.stdout.write(
                            self.style.ERROR(
                                f"Error en la fila {row_number}: Número incorrecto de valores. Verifica la estructura del archivo Excel."
                            )
                        )
                        error_count += 1
                        continue

                    if row_number >= 2:
                        name_split = row[1].strip().split(" ")  # type: ignore

                        first_name = ""
                        last_name = ""
                        for index, name_field in enumerate(name_split):
                            if index < len(name_split) - 2:
                                first_name = f"{first_name} {name_field}"
                            else:
                                last_name = f"{last_name} {name_field}"
                        password = Faker().password(length=12)
                        
                        user = User.objects.create_user(  # type: ignore
                            first_name=first_name.upper().strip(),
                            last_name=last_name.upper().strip(),
                            curp=row[5].upper().strip(),  # type: ignore
                            number_phone=row[4],  # type: ignore
                            address=row[2].strip(),  # type: ignore
                            password=password,
                            email=str(row[9]).lower().strip(),
                        )
                        user.groups.add(group)
                        user.save()

                        teacher = Teacher.objects.create(
                            user=user,
                            budget_code=str(row[6]),
                            academic_level=AcademicLevel.objects.get(text=str(row[7])),
                        )

                        # Añadir el correo y la contraseña al nuevo archivo Excel
                        output_sheet.append([str(row[9]).strip(), password])

                        self.stdout.write(
                            self.style.SUCCESS(f"Creado maestro: {teacher}")
                        )
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f"Error al procesar la fila {row_number}: {str(e)}."
                        )
                    )
                    error_count += 1

            # Guardar el nuevo archivo Excel
            output_workbook.save(output_excel_file_path)
            self.stdout.write(
                self.style.SUCCESS(
                    f"Contraseñas y correos guardados en {output_excel_file_path}"
                )
            )
            self.stdout.write(self.style.SUCCESS("Proceso completado exitosamente."))

        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f"Error: No se encontró el archivo {excel_file_path}.")
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f"Error al procesar el archivo Excel: {str(e)}")
            )
            error_count += 1

        if error_count >= 5:
            self.stdout.write(
                self.style.ERROR("Demasiados errores al procesar el archivo Excel.")
            )
