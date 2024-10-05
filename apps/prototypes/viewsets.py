from datetime import datetime as Datetime

from django.db import IntegrityError, transaction
from django.shortcuts import get_object_or_404

from rest_framework.exceptions import ValidationError
from rest_framework.viewsets import mixins, GenericViewSet
from rest_framework.response import Response
from rest_framework import status

from openpyxl import load_workbook

from apps.cores.serializers import (
    MemberCreatetSerializer,
    PrototypeCreatedSerializer,
    PrototypeDownloadSerializers,
    PrototypeListSerializer,
    PrototypeRetrieveSerializer,
    StudentListSerializer,
    TeacherListSerializer,
)
from apps.prototypes.Templates_excel import (
    PrototypeDonwload,
    PrototypeTemplate,
    StudentTamplete,
    TeacherTemplate,
)
from apps.prototypes.models import Member, Prototype, TeacherRoles
from apps.prototypes.random_teachers import get_teachers_by_modality
from apps.school.models import Student, Teacher
from apps.cores.models import TypeInvestigation

from config.filters import PrototypeDonwloadFilters, PrototypeFilters


def get_index():
    try:
        return (
            int(Prototype.objects.latest("created").registry_number.split("-")[1]) + 1
        )

    except Prototype.DoesNotExist:
        return 1


class PrototypeViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet,
):
    search_fields = ["name"]
    filterset_class = PrototypeFilters

    def get_serializer_class(self):
        if self.action in ["create", "partial_update", "update"]:
            return PrototypeCreatedSerializer

        if self.action in ["retrieve"]:
            return PrototypeRetrieveSerializer

        return PrototypeListSerializer

    def get_queryset(self):
        try:
            if self.request.user.groups.first().name == "Docente":  # type: ignore
                return Prototype.objects.filter(teacher_advisors__teacher_data__user=self.request.user.id)  # type: ignore

            if self.request.user.groups.first().name == "Alumno":  # type: ignore
                return Prototype.objects.filter(
                    members__student__user=self.request.user
                )

            if self.request.user.groups.first().name == "Vinculacion" or request.user.is_staft:  # type: ignore
                return Prototype.objects.all()

            raise ValidationError({"error": "usuario no autorizado"})

        except AttributeError:
            raise ValidationError({"error": "atribute error"})

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        data = request.data
        teacher_advisors = []
        members = []

        if bool(data.get("modality", "")):
            register_number = data.get("modality") + "-" + str(get_index())

            if data.get("teacher_advisors") is None:
                teacher_m, teacher_t = get_teachers_by_modality(data.get("modality"))
                technological = TeacherRoles.objects.create(
                    teacher_data=teacher_t, roles="AT"
                )
                meteorologic = TeacherRoles.objects.create(
                    teacher_data=teacher_m, roles="AM"
                )
                teacher_advisors = [technological.id, meteorologic.id]
            else:
                teacher_advisors = data.get("teacher_advisors", [])

                for teacher in teacher_advisors:
                    if teacher["rol"] in ["AM", "AT"] and len(teacher_advisors) > 0:
                        teacherRole = TeacherRoles.objects.create(
                            teacher_data=Teacher.objects.get(id=teacher["id"]),
                            roles=teacher["rol"],
                        )
                        teacher_advisors.append(teacherRole.id)
                    else:
                        messages = ""
                        if teacher["rol"] == "AM":
                            messages = "Falta el acesor metodologico"

                        if teacher["rol"] == "AT":
                            messages = "Falta el acesor tecnico"

                        raise ValidationError({f"error": messages})

            if len(data.get("members", [])) < 2:
                raise ValidationError(
                    {"error": "Prototipo debe de tener al menos 2 integrantes"}
                )
            for index, student in enumerate(data.get("members")):
                try:
                    stundet = Student.objects.get(id=student)
                    member, create = Member.objects.get_or_create(
                        author=index + 1, student=stundet
                    )
                    members.append(member.id)
                    if not create:
                        raise ValidationError(
                            {
                                "error": f"El alumnos {index + 1}, ya esta assignado a un prototipo"
                            }
                        )

                except Student.DoesNotExist:
                    raise ValidationError({"error": "El alumnos no esta registrado"})

                except IntegrityError:
                    raise ValidationError(
                        {"error": f"El alumno {index + 1}, ya esta registrado."}
                    )

            if len(members) > 4:
                raise ValidationError(
                    {"error": "No se pueden asignar mas de 4 alumnos a un prototipo"}
                )
            data.update(
                {
                    "registry_number": register_number,
                    "teacher_advisors": teacher_advisors,
                    "members": members,
                }
            )

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)

        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )

    # TODO: update modificar multiples de calificaciones de los prototipos

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        data = request.data
        students = data.get("members")
        members = []
        instance_member = [
            member for member in instance.members.all().order_by("author")
        ]

        if len(students) < 2:
            raise ValidationError(
                {"error": "Prototipo debe de tener al menos 2 integrantes"}
            )

        if len(students) > len(instance_member):
            ranged_members = len(students)
        else:
            ranged_members = len(instance_member)

        for index in range(0, ranged_members):
            try:
                if str(instance_member[index].student.id) in students:
                    s_index = students.index(str(instance_member[index].student.id))
                    member = Member.objects.get(student=students[s_index])

                    serializer_member = MemberCreatetSerializer(
                        member,
                        data={"author": index + 1},
                        partial=kwargs.pop("partial", True),
                    )
                    serializer_member.is_valid(raise_exception=True)
                    serializer_member.save()
                    members.append(str(member.id))

                if str(instance_member[index].student.id) not in students:
                    get_object_or_404(Member, id=instance_member[index].id).delete()

            except Member.DoesNotExist:
                serializer_member = MemberCreatetSerializer(
                    data={"author": index + 1, "student": students[index]}
                )
                serializer_member.is_valid(raise_exception=True)
                serializer_member.save()
                members.append(serializer_member.data["id"])  # type: ignore

            except IndexError:
                serializer_member = MemberCreatetSerializer(
                    data={"author": index + 1, "student": students[index]}
                )
                serializer_member.is_valid(raise_exception=True)
                serializer_member.save()
                members.append(serializer_member.data["id"])  # type: ignore

        if len(members) > 4:
            raise ValidationError(
                {"error": "No se pueden asignar mas de 4 alumnos a un prototipo"}
            )

        data.update(
            {
                "members": members,
                "registry_number": f"{data.get('modality')}-{instance.registry_number.split('-')[1]}",  # type: ignore
            }
        )

        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, "_prefetched_objects_cache", None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


class PrototypeDownloadReporterViewsets(PrototypeDonwload):
    queryset = Prototype.objects.all()
    serializer_class = PrototypeDownloadSerializers
    filterset_class = PrototypeDonwloadFilters

    start_of_year = Datetime(Datetime.now().year, 1, 1)
    end_of_year = Datetime(Datetime.now().year, 12, 31)


class StudentTemplateExcelDownload(StudentTamplete):
    serializer_class = StudentListSerializer

    def get_queryset(self):
        return Student.objects.all()


class TeacherTemplateExcelDownload(TeacherTemplate):
    serializer_class = TeacherListSerializer

    def get_queryset(self):
        return Teacher.objects.all()


class PrototypeTemplateExcelDownload(PrototypeTemplate):
    serializer_class = PrototypeDownloadSerializers

    def get_queryset(self):
        return Prototype.objects.all()


def get_modality(modality: str):
    MODALITY = {
        "TECNOLOGICO": "TEC",
        "SOFTWARE": "SW",
        "DIDACTICO": "DC",
        "EMPRENDEDOR_VERDE": "EV",
        "EMPRENDEDOR_SOCIAL": "ES",
        "EMPRENDEDOR_TECNOLOGICO": "ET",
    }
    return MODALITY.get(modality, "")


class UploadPrototypeViewSets(mixins.CreateModelMixin, GenericViewSet):
    serializer_class = PrototypeCreatedSerializer

    def get_queryset(self):
        return Student.objects.all()

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        file_xlx = request.data.get("file")
        members_list_id = []

        if not file_xlx:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        wb = load_workbook(filename=file_xlx, read_only=True)

        for sheet in wb.worksheets:
            # Iterar sobre cada fila de la hoja
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0:
                    continue

                try:
                    for index, student_enumerate in enumerate(range(8, 15, 2)):
                        if row[student_enumerate]:
                            member = Member.objects.create(
                                author=index + 1,
                                student=Student.objects.get(
                                    user__email=row[student_enumerate]
                                ),
                            )

                            members_list_id.append(member.id)
                except IntegrityError:
                    continue

                advisors_technological_id = TeacherRoles.objects.create(
                    roles="AT",
                    teacher_data=Teacher.objects.get(user__email=row[5]),
                ).id

                advisors_meteorologic_id = TeacherRoles.objects.create(
                    roles="MT",
                    teacher_data=Teacher.objects.get(user__email=row[6]),
                ).id

                serializer = self.get_serializer(
                    data={
                        "name": row[0],
                        "registry_number": row[1],
                        "modality": get_modality(row[2]),
                        "type_investigation": TypeInvestigation.objects.get(
                            text=row[3].strip()
                        ).id,
                        "members": members_list_id,
                        "teacher_methods": Teacher.objects.get(user__email=row[4]).id,
                        "teacher_advisors": [
                            advisors_technological_id,
                            advisors_meteorologic_id,
                        ],
                    }
                )

                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)

        try:
            headers = self.get_success_headers(serializer.data)
        except UnboundLocalError:
            raise ValidationError({"prototype": "este archivo ya esta importado"})

        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )
