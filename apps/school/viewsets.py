from django.db import transaction
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group

from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
from rest_framework.viewsets import mixins, GenericViewSet
from rest_framework.response import Response
from rest_framework import status

from apps.cores.models import AcademicGroup
from apps.cores.serializers import (
    StudentCreateSerializer,
    StudentListSerializer,
    StudentRetrieveSerializer,
    TeacherCreateSerializer,
    TeacherListSerializer,
    TeacherRetrieveSerializer,
)
from apps.school.models import Student, Teacher
from apps.users.serializers import UserCreateSerializer, UserUpdateSerializer
from config.filters import StudentsFilters, TeachersFilters

User = get_user_model()


class StudentViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet,
):
    search_fields = ["user__email", "user__turn"]
    ordering_fields = ["group"]
    filterset_class = StudentsFilters

    def get_queryset(self):
        return Student.objects.all()

    def get_serializer_class(self):
        print(self.action)
        if self.action in ["create", "partial_update", "update"]:
            return StudentCreateSerializer

        if self.action in ["retrieve"]:
            return StudentRetrieveSerializer

        return StudentListSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        data = request.data

        try:
            group = AcademicGroup.objects.get(id=data.get("group", None)).id
        except AcademicGroup.DoesNotExist:
            group = None

        data.update({"group": group})

        serializer_user = UserCreateSerializer(
            data={
                "first_name": data.get("user").get("first_name", None),
                "last_name": data.get("user").get("last_name", None),
                "curp": data.get("user").get("curp", None),
                "email": data.get("user").get("email", None),
                "number_phone": data.get("user").get("number_phone", None),
                "address": data.get("user").get("address", None),
                "groups": [Group.objects.get(name="Alumno").pk],
                "turn": data.get("user").get("turn", None),
            }
        )
        serializer_user.is_valid(raise_exception=True)
        self.perform_create(serializer_user)

        data.update(
            {
                "user": serializer_user.data.get("id", None),  # type: ignore
            }
        )

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        headers = self.get_success_headers(serializer.data)

        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        data = request.data
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        if data.get("user", None) != None:

            serializer_user = UserUpdateSerializer(
                instance=User.objects.get(id=instance.user_id),
                data=data.get("user"),
                partial=partial,
            )
            serializer_user.is_valid(raise_exception=True)
            self.perform_update(serializer_user)

            data.update({"user": serializer_user.data.get("id", None)})  # type: ignore

        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, "_prefetched_objects_cache", None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


class TeacherViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    GenericViewSet,
):
    search_fields = ["user__email", "user__first_name", "user__last_name", "user__curp"]
    ordering_fields = [
        "user__first_name",
        "user__last_name",
        "user__email",
        "budget_code, user__turn",
    ]

    filterset_class = TeachersFilters

    def get_queryset(self):
        return Teacher.objects.all()

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return TeacherCreateSerializer

        if self.action in ["retrieve"]:
            return TeacherRetrieveSerializer

        return TeacherListSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        data = request.data

        if not data.get("user", None):
            raise ValidationError({"user": "Este campo es requerido"})

        serializer_user = UserCreateSerializer(
            data={
                "first_name": data.get("user").get("first_name", None),
                "last_name": data.get("user").get("last_name", None),
                "curp": data.get("user").get("curp", None),
                "email": data.get("user").get("email", None),
                "number_phone": data.get("user").get("number_phone", None),
                "address": data.get("user").get("address", None),
                "groups": [Group.objects.get(name="Docente").pk],
                "turn": data.get("user").get("turn", None),
            }
        )
        serializer_user.is_valid(raise_exception=True)
        self.perform_create(serializer_user)

        data.update({"user": serializer_user.data.get("id", None)})

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        data = request.data
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        if data.get("user", None) != None:

            serializer_user = UserUpdateSerializer(
                instance=User.objects.get(id=instance.user_id),
                data=data.get("user"),
                partial=partial,
            )
            serializer_user.is_valid(raise_exception=True)
            self.perform_update(serializer_user)

            data.update({"user": serializer_user.data.get("id", None)})  # type: ignore

        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        if getattr(instance, "_prefetched_objects_cache", None):
            # If 'prefetch_related' has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return Response(serializer.data)


SPECIALITY = {
    "TECNICO EN PROGRAMACION": "TEP",
    "TECNICO EN CONTABILIDAD": "TEC",
    "TECNICO EN SECRETARIADO EJECUTIVO BILINGUE": "TESEB",
    "TECNICO EN CIENCIA DE DATOS E INFORMACION": "TECDEI",
}


class UploadStudentViewSets(mixins.CreateModelMixin, GenericViewSet):
    serializer_class = StudentCreateSerializer

    def get_queryset(self):
        return Student.objects.all()

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        file_xlx = request.data.get("file")

        if not file_xlx:
            return Response(
                {"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        wb = load_workbook(filename=file_xlx, read_only=True)

        for sheet in wb.worksheets:
            # Iterar sobre cada fila de la hoja
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index == 0:
                    continue

                if row[6] == "VESPERTINO":
                    turn = "T/V"

                elif row[6] == "MATUTINO":
                    turn = "T/M"

                if row[4] != None:
                    if User.objects.get(email=row[4]).email == row[4]: 
                        continue

                    serializer_user = UserCreateSerializer(
                        data={
                            "first_name": row[0],
                            "last_name": row[1],
                            "curp": row[2],
                            "address": row[3],
                            "email": row[4],
                            "number_phone": row[5],
                            "groups": [Group.objects.get(name="Alumno").pk],
                            "turn": turn,
                        }
                    )

                    serializer_user.is_valid(raise_exception=True)
                    self.perform_create(serializer_user)

                    serializer = self.get_serializer(
                        data={
                            "user": serializer_user.data.get("id", None),
                            "school_control_number": row[7],
                            "group": AcademicGroup.objects.get(text=row[8]).id,
                            "specialty": SPECIALITY.get(row[9]),
                        }
                    )
                    serializer.is_valid(raise_exception=True)
                    self.perform_create(serializer)

        try:
            headers = self.get_success_headers(serializer.data)
        except UnboundLocalError:
            raise ValidationError({"students": "este archivo ya esta importado"})

        return Response(
            serializer.data, status=status.HTTP_201_CREATED, headers=headers
        )
